#!/usr/bin/env python3
"""
Export Gantt Chart to Excel

This script takes the processed task data and exports it to an Excel file
with formatting to represent the Gantt chart visually.
"""

import pandas as pd
import argparse
import os
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color
from openpyxl.utils import get_column_letter
import json

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Export Gantt chart data to Excel.')
    parser.add_argument('--input', '-i', required=True, help='Path to the input Excel file with task data')
    parser.add_argument('--output', '-o', default='gantt_chart_export.xlsx', help='Path to save the output Excel file')
    return parser.parse_args()

def read_excel_data(file_path):
    """Read data from the Excel file."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Input file not found: {file_path}")
    
    try:
        df = pd.read_excel(file_path)
        
        # Map expected column names to actual column names
        expected_columns = {
            'Task': 'Task',
            'Task 1': 'Task 1',
            'Business Driver': 'Business Driver',
            'Resource': 'Resources',  # Note: 'Resources' in input file
            'Data': 'Data',
            'Location': 'Location'
        }
        
        # Check if the required columns exist
        missing_columns = []
        for expected_col in ['Task', 'Resources', 'Location', 'Business Driver']:
            if expected_col not in df.columns:
                missing_columns.append(expected_col)
        
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
        
        # Check for month columns - using full month names
        full_month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                           'July', 'August', 'September', 'October', 'November', 'December']
        
        month_columns = [col for col in df.columns if col in full_month_names]
        if not month_columns:
            raise ValueError("No month columns found (January-December)")
        
        return df
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")

def process_data_for_gantt(df):
    """Process the Excel data into a format suitable for Gantt chart."""
    current_year = datetime.now().year
    
    # Create a list to store task data
    tasks = []
    
    # Month to number mapping using full month names
    month_to_num = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
        'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12
    }
    
    # Process each row
    for _, row in df.iterrows():
        task_name = row['Task']
        resource = row['Resources']  # Note: using 'Resources' instead of 'Resource'
        location = row['Location']
        driver = row.get('Business Driver', '')  # Use get to handle missing column
        task1 = row.get('Task 1', '')  # Get Task 1 value
        data = row.get('Data', '')  # Get Data value
        
        # Find start and end months
        start_month = None
        end_month = None
        duration = 0
        
        for month in month_to_num.keys():
            if month in df.columns and pd.notna(row[month]) and row[month]:
                if start_month is None:
                    start_month = month_to_num[month]
                end_month = month_to_num[month]
                duration += 1
        
        # Skip if no duration
        if start_month is None or end_month is None:
            continue
        
        # Create start and end dates
        start_date = datetime(current_year, start_month, 1)
        # End date is the last day of the end month
        if end_month == 12:
            end_date = datetime(current_year, 12, 31)
        else:
            end_date = datetime(current_year, end_month + 1, 1) - timedelta(days=1)
        
        # Use Task 1 as the primary identifier, fallback to Task if Task 1 is empty
        display_name = task1 if pd.notna(task1) and task1 else task_name
        
        # Create a combined key for Resource, Driver, Location
        resource_key = f"{resource} | {driver} | {location}"
        
        # Add task to the list
        tasks.append({
            'Task': task_name,
            'Task 1': task1,
            'Display_Name': display_name,  # New field for display purposes
            'Resource': resource,
            'Location': location,
            'Driver': driver,
            'Data': data,
            'ResourceKey': resource_key,  # Combined key for grouping
            'Start': start_date,
            'Finish': end_date,
            'Duration': duration,
            'Start_Month': start_month,
            'End_Month': end_month
        })
    
    return pd.DataFrame(tasks)

def calculate_resource_percentage(df_tasks, task):
    """Calculate the percentage of unique resource months spent on a task."""
    resource = task['Resource']
    
    # Get all tasks for this resource
    resource_tasks = df_tasks[df_tasks['Resource'] == resource]
    
    # Calculate total unique resource-months
    total_resource_months = len(set([
        (t['Resource'], m) 
        for _, t in resource_tasks.iterrows() 
        for m in range(t['Start_Month'], t['End_Month'] + 1)
    ]))
    
    # Calculate unique resource-months for this task
    task_resource_months = len(set([
        (resource, m) 
        for m in range(task['Start_Month'], task['End_Month'] + 1)
    ]))
    
    # Calculate percentage
    return (task_resource_months / total_resource_months) * 100 if total_resource_months > 0 else 0


def calculate_task1_percentages(df_tasks, display_color_map):
    """Calculate percentages of unique resource months for each Task 1 (Display_Name)."""
    task1_resource_months = {}
    total_resource_months_by_display = {}
    task1_percentages = {}
    
    # First, calculate total unique resource-months for each display name
    for display_name in sorted(display_color_map.keys()):
        tasks_with_display = df_tasks[df_tasks['Display_Name'] == display_name]
        
        # Get all resources used by this display name
        resources = tasks_with_display['Resource'].unique()
        
        # For each resource, get all tasks with that resource
        all_resource_months = set()
        task_resource_months = set()
        
        for resource in resources:
            resource_tasks = df_tasks[df_tasks['Resource'] == resource]
            
            # Add all resource-month pairs to the set
            for _, t in resource_tasks.iterrows():
                for m in range(t['Start_Month'], t['End_Month'] + 1):
                    all_resource_months.add((t['Resource'], m))
            
            # Add resource-month pairs for this display name to the set
            for _, t in tasks_with_display[tasks_with_display['Resource'] == resource].iterrows():
                for m in range(t['Start_Month'], t['End_Month'] + 1):
                    task_resource_months.add((t['Resource'], m))
        
        # Store the counts
        task1_resource_months[display_name] = len(task_resource_months)
        total_resource_months_by_display[display_name] = len(all_resource_months)
        
        # Calculate percentage
        task_months = task1_resource_months.get(display_name, 0)
        total_months = total_resource_months_by_display.get(display_name, 0)
        task1_percentages[display_name] = (task_months / total_months) * 100 if total_months > 0 else 0
    
    return task1_percentages


def create_excel_gantt(df_tasks, output_file):
    """Create an Excel file with Gantt chart representation."""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"
    
    # Sort tasks by Resource, Driver, Location, and Start date
    df_tasks = df_tasks.sort_values(by=['Resource', 'Driver', 'Location', 'Start'])
    
    # Define colors directly based on Display_Name (which is Task 1 or Task if Task 1 is empty)
    display_names = df_tasks['Display_Name'].unique()
    
    colors = [
        "1F77B4", "FF7F0E", "2CA02C", "D62728", "9467BD", 
        "8C564B", "E377C2", "7F7F7F", "BCBD22", "17BECF"
    ]
    
    # Create color map based on Display_Name (Task 1 values)
    display_color_map = {}
    for i, name in enumerate(sorted(display_names)):
        display_color_map[name] = colors[i % len(colors)]
    
    # Month headers - using full month names
    month_headers = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    
    # Set up headers
    headers = ["Resource", "Driver", "Location"]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Write month headers
    for col, month in enumerate(month_headers, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.value = month
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Group by ResourceKey (Resource | Driver | Location)
    row_idx = 2
    current_resource_key = None
    
    # Process each resource group
    for resource_key, group in df_tasks.groupby('ResourceKey'):
        resource, driver, location = resource_key.split(' | ')
        
        # No blank rows between resource groups
        
        # Write the resource, driver, location in the first columns
        ws.cell(row=row_idx, column=1).value = resource
        ws.cell(row=row_idx, column=2).value = driver
        ws.cell(row=row_idx, column=3).value = location
        
        # Add background color to the resource info cells
        for col in range(1, 4):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            cell.font = Font(bold=True)
        
        # For each task in this resource group, add bars in the month columns
        for _, task in group.iterrows():
            # For each month in the task's duration
            for month_idx in range(task['Start_Month'], task['End_Month'] + 1):
                col_idx = month_idx + len(headers)  # Adjust column index for month
                
                # Add the task to the cell
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Color based on Display_Name (Task 1 value)
                cell.fill = PatternFill(start_color=display_color_map[task['Display_Name']], 
                                       end_color=display_color_map[task['Display_Name']], 
                                       fill_type="solid")
                
                # Add Display_Name (Task 1) with percentage to the first cell of the bar
                if month_idx == task['Start_Month']:
                    # Calculate percentage of unique resource months for this task
                    percentage = calculate_resource_percentage(df_tasks, task)
                    
                    # Format as "Task 1 Name (XX%)"
                    display_text = f"{task['Display_Name']} ({percentage:.0f}%)"
                    
                    cell.value = display_text
                    cell.alignment = Alignment(horizontal='left')
                    cell.font = Font(color="FFFFFF", bold=True)
        
        row_idx += 1
    
    # Add a summary sheet
    ws_summary = wb.create_sheet(title="Resource Summary")
    
    # Create a summary table by resource, driver, location
    # Make sure to handle empty values properly
    resource_summary = df_tasks.groupby(['Resource', 'Driver', 'Location']).agg(
        Tasks=('Task', lambda x: ', '.join(sorted(set(x)))),
        Task1=('Task 1', lambda x: ', '.join(sorted(set([str(i) for i in x if pd.notna(i) and i != ''])))),
        Data=('Data', lambda x: ', '.join(sorted(set([str(i) for i in x if pd.notna(i) and i != ''])))),
        Task_Count=('Task', 'count'),
        Total_Duration=('Duration', 'sum'),
        Months=('Task', lambda x: ', '.join([month_headers[m-1] for m in sorted(set(
            [month for task_idx, task in df_tasks[df_tasks['Task'].isin(x)].iterrows() 
             for month in range(task['Start_Month'], task['End_Month']+1)]))]))
    ).reset_index()
    
    # Replace 'nan' strings with empty strings
    for col in ['Task1', 'Data']:
        resource_summary[col] = resource_summary[col].str.replace('nan', '').str.strip(', ')
        # Also handle cases where the string starts or ends with a comma
        resource_summary[col] = resource_summary[col].str.strip(', ')
    
    # Write summary headers
    summary_headers = ["Resource", "Driver", "Location", "Tasks", "Task 1", "Data", "Task Count", "Total Duration (months)", "Months"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Write summary data
    for i, (_, row) in enumerate(resource_summary.iterrows(), 2):
        ws_summary.cell(row=i, column=1).value = row['Resource']
        ws_summary.cell(row=i, column=2).value = row['Driver']
        ws_summary.cell(row=i, column=3).value = row['Location']
        ws_summary.cell(row=i, column=4).value = row['Tasks']
        ws_summary.cell(row=i, column=5).value = row['Task1']
        ws_summary.cell(row=i, column=6).value = row['Data']
        ws_summary.cell(row=i, column=7).value = row['Task_Count']
        ws_summary.cell(row=i, column=8).value = row['Total_Duration']
        ws_summary.cell(row=i, column=9).value = row['Months']
        
        # Add background color to distinguish rows
        for col in range(1, 10):
            cell = ws_summary.cell(row=i, column=col)
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Add a task legend sheet for Task 1 values
    ws_legend = wb.create_sheet(title="Task 1 Legend")
    
    # Write legend headers
    legend_headers = ["Task 1", "Color"]
    for col, header in enumerate(legend_headers, 1):
        cell = ws_legend.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Calculate percentage of unique resource months for each Task 1
    task1_percentages = calculate_task1_percentages(df_tasks, display_color_map)
    
    # Write Display_Name (Task 1) colors with percentage
    for i, display_name in enumerate(sorted(display_color_map.keys()), 2):
        # Get percentage from calculated values
        percentage = task1_percentages.get(display_name, 0)
        
        # Format as "Task 1 Name (XX%)"
        display_text = f"{display_name} ({percentage:.0f}%)"
        
        ws_legend.cell(row=i, column=1).value = display_text
        
        # Color cell based on Display_Name
        color_cell = ws_legend.cell(row=i, column=2)
        color_cell.fill = PatternFill(start_color=display_color_map[display_name], 
                                     end_color=display_color_map[display_name], 
                                     fill_type="solid")
    
    # Auto-adjust column widths
    for ws_name in [ws, ws_summary]:
        for col in range(1, ws_name.max_column + 1):
            max_length = 0
            for row in range(1, ws_name.max_row + 1):
                cell = ws_name.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 10)
            ws_name.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel Gantt chart saved to {output_file}")

def main():
    """Main function to run the script."""
    args = parse_arguments()
    
    try:
        # Read data from Excel
        print(f"Reading data from {args.input}...")
        df = read_excel_data(args.input)
        
        # Process data for Gantt chart
        print("Processing data...")
        df_tasks = process_data_for_gantt(df)
        
        # Create Excel Gantt chart
        print(f"Creating Excel Gantt chart and saving to {args.output}...")
        create_excel_gantt(df_tasks, args.output)
        
        print(f"\nGantt chart successfully created and saved to {args.output}")
        print("\nThe Excel file contains:")
        print("1. Gantt Chart - Visual representation of tasks by Resource, Driver, Location")
        print("2. Resource Summary - Details of tasks grouped by Resource, Driver, Location")
        print("3. Task Legend - Color coding for each task")
    except Exception as e:
        print(f"Error: {str(e)}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())
